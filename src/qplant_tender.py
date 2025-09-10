from __future__ import annotations

"""Utilities for inserting QPLANT tender milestones into the master Excel file."""

from pathlib import Path
from typing import Dict, Iterable

import pandas as pd

# Baseline purchase order signature for QPLANT
BASELINE_PO = pd.Timestamp("2027-06-01")

# Scenario definitions: name -> (PO date, fill colour)
SCENARIOS: Dict[str, tuple[pd.Timestamp, str]] = {
    "baseline": (BASELINE_PO, "C6EFCE"),  # green
    "+2m": (BASELINE_PO + pd.DateOffset(months=2), "FFD966"),  # orange
    "+4m": (BASELINE_PO + pd.DateOffset(months=4), "F4CCCC"),  # red
}

# Addendum II milestones relative to PO signature
ADDENDUM_OFFSETS = {
    "M16 Conceptual": 16,
    "M18 Detailed": 18,
    "M23 Manufacturing": 23,
    "M44 Provisional Acceptance": 44,
}

# Baseline tender milestones
MILESTONES = {
    "Launch": pd.Timestamp("2025-11-01"),
    "Eval": pd.Timestamp("2026-03-01"),
    "Neg": pd.Timestamp("2026-05-01"),
}


def _ensure_qplant_in_list(df: pd.DataFrame) -> pd.DataFrame:
    """Return *df* with a QPLANT row appended if missing."""
    col = df.columns[0]
    if "QPLANT" not in df[col].astype(str).str.upper().values:
        df = pd.concat([df, pd.DataFrame({col: ["QPLANT"]})], ignore_index=True)
    return df


def _build_monitoring_rows() -> pd.DataFrame:
    """Return monitoring rows for all scenarios."""
    rows = []
    for scenario, (po_date, _colour) in SCENARIOS.items():
        row = {
            "Element": "Cryoplant",
            "Description": "Helium refrigeration tender",
            "Mnemonic": "QPL",
            "SPOC": "",
            "Dossier": "",
            "Launch": MILESTONES["Launch"].date(),
            "Eval": MILESTONES["Eval"].date(),
            "Neg": MILESTONES["Neg"].date(),
            "PO": po_date.date(),
            "Scenario": scenario,
        }
        for name, months in ADDENDUM_OFFSETS.items():
            row[name] = (po_date + pd.DateOffset(months=months)).date()
        rows.append(row)
    return pd.DataFrame(rows)


def update_qplant_schedule(workbook: Path) -> Path:
    """Insert QPLANT data into *workbook* and export coloured schedule.

    Args:
        workbook: Path to ``Detailing tender Process of ATS Tenders.xlsx``.

    Returns:
        Path to the generated ``tender_schedule_colored.xlsx`` file.
    """
    xls = pd.ExcelFile(workbook)
    list_df = _ensure_qplant_in_list(xls.parse("List"))
    monitor_df = xls.parse("Detailed Monitoring ATS Tenders")

    monitor_df = pd.concat([monitor_df, _build_monitoring_rows()], ignore_index=True)

    with pd.ExcelWriter(workbook, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        list_df.to_excel(writer, sheet_name="List", index=False)
        monitor_df.to_excel(writer, sheet_name="Detailed Monitoring ATS Tenders", index=False)

    colour_path = workbook.with_name("tender_schedule_colored.xlsx")
    with pd.ExcelWriter(colour_path, engine="openpyxl") as writer:
        monitor_df.to_excel(writer, sheet_name="QPLANT", index=False)
        from openpyxl.styles import PatternFill  # Imported here to allow optional dependency

        ws = writer.sheets["QPLANT"]
        for row_idx, scenario in enumerate(monitor_df["Scenario"], start=2):
            colour = SCENARIOS[scenario][1]
            fill = PatternFill(start_color=colour, end_color=colour, fill_type="solid")
            for col_idx in range(1, monitor_df.shape[1] + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill
    return colour_path


def main(args: Iterable[str] | None = None) -> int:
    """CLI entry point."""
    import argparse

    parser = argparse.ArgumentParser(description="Update QPLANT tender schedule")
    parser.add_argument("workbook", type=Path, help="Path to the tender workbook")
    parsed = parser.parse_args(list(args) if args is not None else None)

    path = update_qplant_schedule(parsed.workbook)
    print(f"Colored schedule written to {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
